"""
DART 재무분석 웹앱 — FastAPI 백엔드 (포트 8002)
v8 로직 완전 이식: CODE_MAP/NAME_MAP 계정 분류 + Claude CLI 종합 의견
"""
from __future__ import annotations

import io
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from typing import Optional

import openpyxl
import requests as http_req
import xml.etree.ElementTree as ET
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

# ── 상수 ──────────────────────────────────────────────────────────────────────
# 환경변수 우선, 없으면 하드코딩 fallback (로컬 개발용)
DART_KEY    = os.environ.get("DART_API_KEY", "1592942f6fb95b1c8cd9fc5e81479ae025fc01e4")
ANTHROPIC_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
CORP_URL    = "https://opendart.fss.or.kr/api/corpCode.xml"
FIN_URL     = "https://opendart.fss.or.kr/api/fnlttSinglAcntAll.json"
YEARS       = [2023, 2024, 2025]
REPRT       = "11011"
CLI_TIMEOUT = 120

_HERE      = os.path.dirname(os.path.abspath(__file__))
# Vercel 서버리스: /tmp 사용, 로컬: api/ 폴더 사용
_TMP       = "/tmp" if os.path.isdir("/tmp") else _HERE
OUTPUT_DIR = os.path.join(_TMP, "dart_output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 번들된 CORPCODE.xml 경로 (배포 시 런타임 다운로드 불필요)
_BUNDLED_XML = os.path.join(_HERE, "CORPCODE.xml")

# ── 계정 매핑 ─────────────────────────────────────────────────────────────────
TARGET_KEYS = ["매출액", "영업이익", "당기순이익", "자산총계", "부채총계", "자본총계"]
_IS_KEYS    = frozenset(["매출액", "영업이익", "당기순이익"])
_BS_KEYS    = frozenset(["자산총계", "부채총계", "자본총계"])

CODE_MAP: dict[str, str] = {
    "ifrs-full_Revenue":                                 "매출액",
    "ifrs-full_RevenueFromContractsWithCustomers":       "매출액",
    "dart_Revenue":                                      "매출액",
    "dart_OperatingIncomeLoss":                          "영업이익",
    "ifrs-full_ProfitLossFromOperatingActivities":       "영업이익",
    "ifrs-full_ProfitLoss":                              "당기순이익",
    "ifrs-full_ProfitLossAttributableToOwnersOfParent":  "당기순이익",
    "dart_ProfitLossAttributableToOwnersOfParentEntity": "당기순이익",
    "ifrs-full_Assets":                                  "자산총계",
    "ifrs-full_Liabilities":                             "부채총계",
    "ifrs-full_Equity":                                  "자본총계",
    "ifrs-full_EquityAttributableToOwnersOfParent":      "자본총계",
}

NAME_MAP: dict[str, str] = {
    "매출액": "매출액", "수익(매출액)": "매출액", "매출수익": "매출액",
    "영업수익": "매출액", "도급공사수익": "매출액", "공사수익": "매출액",
    "용역수익": "매출액", "보험료수익": "매출액", "이자수익": "매출액",
    "순이자수익": "매출액", "수수료수익": "매출액", "서비스수익": "매출액",
    "운임수익": "매출액", "임대수익": "매출액", "보험영업수익": "매출액",
    "영업이익": "영업이익", "영업이익(손실)": "영업이익", "영업손익": "영업이익",
    "당기순이익": "당기순이익", "당기순이익(손실)": "당기순이익",
    "당기순손익": "당기순이익", "분기순이익": "당기순이익", "반기순이익": "당기순이익",
    "자산총계": "자산총계", "부채총계": "부채총계", "자본총계": "자본총계",
}

# ── 기업코드 캐시 ──────────────────────────────────────────────────────────────
_corps_cache: list | None = None
_corps_lock  = threading.Lock()

# ── DART API 함수 ──────────────────────────────────────────────────────────────
def load_corp_codes() -> list[dict]:
    global _corps_cache
    with _corps_lock:
        if _corps_cache:
            return _corps_cache

        # 1순위: 프로젝트에 번들된 XML (Vercel 배포 시 런타임 다운로드 불필요)
        # 2순위: /tmp 캐시 (이전 호출에서 저장한 경우)
        # 3순위: DART API에서 실시간 다운로드
        for xml_path in [_BUNDLED_XML, os.path.join(_TMP, "CORPCODE.xml")]:
            if os.path.exists(xml_path):
                root = ET.parse(xml_path).getroot()
                break
        else:
            r = http_req.get(CORP_URL, params={"crtfc_key": DART_KEY}, timeout=30)
            r.raise_for_status()
            zf   = zipfile.ZipFile(io.BytesIO(r.content))
            root = ET.fromstring(zf.read("CORPCODE.xml"))
            # /tmp에 캐시 저장 (다음 호출에서 재사용)
            try:
                with open(os.path.join(_TMP, "CORPCODE.xml"), "wb") as f:
                    f.write(ET.tostring(root, encoding="utf-8", xml_declaration=True))
            except Exception:
                pass

        _corps_cache = [
            {k: (item.findtext(k) or "").strip()
             for k in ("corp_code", "corp_name", "stock_code")}
            for item in root.findall(".//list")
        ]
        return _corps_cache


def search_corp(corps: list, name: str) -> list[dict]:
    exact = [c for c in corps if c["corp_name"] == name]
    return exact or [c for c in corps if name in c["corp_name"]]


def fetch_fin(corp_code: str, year: int, fs_div: str = "CFS", _retry: int = 3) -> list | None:
    params = dict(crtfc_key=DART_KEY, corp_code=corp_code,
                  bsns_year=year, reprt_code=REPRT, fs_div=fs_div)
    for attempt in range(_retry):
        try:
            r = http_req.get(FIN_URL, params=params, timeout=60)
            r.raise_for_status()
            if not r.text.strip():
                raise ValueError("빈 응답")
            d = r.json()
            if d.get("status") != "000":
                return fetch_fin(corp_code, year, "OFS") if fs_div == "CFS" else None
            return d.get("list", [])
        except (ValueError, Exception) as e:
            if attempt < _retry - 1:
                time.sleep(2 ** attempt)
                continue
            if fs_div == "CFS":
                return fetch_fin(corp_code, year, "OFS")
            return None
    return None


def parse_items(items: list) -> dict:
    """CODE_MAP(account_id) 우선, NAME_MAP(account_nm) 폴백"""
    result: dict[str, int | None] = {}
    is_items = [x for x in items if x.get("sj_div") in ("IS", "CIS")]
    bs_items = [x for x in items if x.get("sj_div") == "BS"]

    for std_key in TARGET_KEYS:
        scope = (bs_items if bs_items else items) if std_key in _BS_KEYS \
                else (is_items if is_items else items)
        found = None
        for item in scope:
            if CODE_MAP.get(item.get("account_id", "")) == std_key:
                found = item; break
        if not found:
            for item in scope:
                if NAME_MAP.get(item.get("account_nm", "").strip()) == std_key:
                    found = item; break
        if found:
            raw = (found.get("thstrm_amount") or "0").replace(",", "").strip()
            try:
                result[std_key] = int(raw)
            except Exception:
                result[std_key] = 0
        else:
            result[std_key] = None
    return result


# ── 계산 유틸 ─────────────────────────────────────────────────────────────────
def calc_ratios(d: dict) -> dict:
    rev, op = d.get("매출액"), d.get("영업이익")
    net, liab, eq = d.get("당기순이익"), d.get("부채총계"), d.get("자본총계")
    assets = d.get("자산총계")
    out = {}
    if rev:
        if op  is not None: out["영업이익률"] = round(op  / rev * 100, 2)
        if net is not None: out["순이익률"]   = round(net / rev * 100, 2)
    if eq and eq != 0 and liab is not None:
        out["부채비율"] = round(liab / eq * 100, 2)
    if eq and eq != 0 and net is not None:
        out["ROE"] = round(net / eq * 100, 2)
    if net is not None and assets and assets != 0:
        out["ROA"] = round(net / assets * 100, 2)          # ← ROA 추가
    if eq is not None and assets and assets != 0:
        out["자기자본비율"] = round(eq / assets * 100, 2)
    return out


def yoy(cur, prv) -> float | None:
    if prv and prv != 0 and cur is not None:
        return round((cur - prv) / abs(prv) * 100, 2)
    return None


def fmt_amt(v) -> str:
    return f"{v/1e8:,.1f}" if v is not None else "N/A"

def fmt_pct(v) -> str:
    return f"{v:.1f}%" if v is not None else "N/A"

def fmt_yoy(v) -> str:
    if v is None: return ""
    return f"▲{v:.1f}%" if v >= 0 else f"▼{abs(v):.1f}%"


# ── 병렬 데이터 수집 ──────────────────────────────────────────────────────────
def fetch_company_data_parallel(corps: list[dict]) -> dict[str, dict[int, dict | None]]:
    """3회사 × 3년 = 최대 9개 DART API 동시 호출 (ThreadPoolExecutor)"""
    results: dict[str, dict] = {c["corp_code"]: {} for c in corps}

    def _fetch(corp_code: str, year: int):
        try:
            items = fetch_fin(corp_code, year)
            return corp_code, year, parse_items(items) if items else None
        except Exception:
            return corp_code, year, None

    with ThreadPoolExecutor(max_workers=9) as executor:
        futures = [
            executor.submit(_fetch, c["corp_code"], yr)
            for c in corps for yr in YEARS
        ]
        for f in as_completed(futures):
            corp_code, year, data = f.result()
            results[corp_code][year] = data
    return results


def build_company_result(corp: dict, fin: dict[int, dict | None]) -> dict:
    """재무 데이터에서 ratios / yoy 계산하여 결과 딕셔너리 반환"""
    ratios: dict[int, dict] = {}
    yoy_data: dict[int, dict] = {}

    for yr in YEARS:
        d = fin.get(yr)
        ratios[yr] = calc_ratios(d) if d else {}

    for i in range(1, len(YEARS)):
        py, cy = YEARS[i - 1], YEARS[i]
        pd_, cd = fin.get(py) or {}, fin.get(cy) or {}
        yoy_data[cy] = {
            k: yoy(cd.get(k), pd_.get(k))
            for k in ["매출액", "영업이익", "당기순이익"]
        }
        # 성장률을 ratios dict에 합산 (프론트 RatioTable이 ratios만 바라봄)
        if yoy_data[cy].get("매출액") is not None:
            ratios[cy]["매출성장률"] = yoy_data[cy]["매출액"]
        if yoy_data[cy].get("영업이익") is not None:
            ratios[cy]["영업이익성장률"] = yoy_data[cy]["영업이익"]

    return {
        "corp":   corp,
        "fin":    {str(yr): fin.get(yr) for yr in YEARS},
        "ratios": {str(yr): ratios.get(yr, {}) for yr in YEARS},
        "yoy":    {str(yr): yoy_data.get(yr, {}) for yr in YEARS},
    }


# ── Claude CLI ────────────────────────────────────────────────────────────────
def _find_claude_cmd() -> list[str]:
    exe = shutil.which("claude")
    if exe:
        if sys.platform == "win32" and exe.lower().endswith(".cmd"):
            return ["cmd", "/c", exe]
        return [exe]
    if sys.platform == "win32":
        for base in [
            os.path.expandvars(r"%APPDATA%\npm"),
            os.path.expanduser(r"~\AppData\Roaming\npm"),
        ]:
            for name in ("claude.cmd", "claude.ps1", "claude"):
                path = os.path.join(base, name)
                if os.path.isfile(path):
                    if name.endswith(".cmd"):
                        return ["cmd", "/c", path]
                    if name.endswith(".ps1"):
                        return ["powershell", "-ExecutionPolicy", "Bypass", "-File", path]
                    return [path]
    return []


def build_cli_prompt(corp_name: str, fin: dict) -> str:
    valid = [yr for yr in YEARS if fin.get(str(yr)) or fin.get(yr)]
    fin_json: dict = {"회사명": corp_name, "분석연도": valid, "연도별데이터": {}}

    for yr in valid:
        d = fin.get(str(yr)) or fin.get(yr) or {}
        r = calc_ratios(d)
        assets = d.get("자산총계")
        eq     = d.get("자본총계")
        er     = round(eq / assets * 100, 2) if eq and assets else None
        fin_json["연도별데이터"][str(yr)] = {
            "매출액":       fmt_amt(d.get("매출액")),
            "영업이익":     fmt_amt(d.get("영업이익")),
            "당기순이익":   fmt_amt(d.get("당기순이익")),
            "자산총계":     fmt_amt(d.get("자산총계")),
            "부채총계":     fmt_amt(d.get("부채총계")),
            "자본총계":     fmt_amt(d.get("자본총계")),
            "영업이익률":   fmt_pct(r.get("영업이익률")),
            "순이익률":     fmt_pct(r.get("순이익률")),
            "부채비율":     fmt_pct(r.get("부채비율")),
            "ROE":          fmt_pct(r.get("ROE")),
            "자기자본비율": fmt_pct(er),
        }

    if len(valid) >= 2:
        yoy_rows: dict = {}
        for i in range(1, len(valid)):
            py, cy = valid[i - 1], valid[i]
            pd_ = fin.get(str(py)) or fin.get(py) or {}
            cd  = fin.get(str(cy)) or fin.get(cy) or {}
            for key in ["매출액", "영업이익", "당기순이익"]:
                v = yoy(cd.get(key), pd_.get(key))
                if v is not None:
                    yoy_rows[f"{key}_{py}→{cy}"] = fmt_yoy(v)
        fin_json["YoY증감률"] = yoy_rows

    fin_str = json.dumps(fin_json, ensure_ascii=False, indent=2)
    return (
        "당신은 15년 경력의 전문 재무 애널리스트입니다.\n"
        "아래 JSON 재무 데이터를 분석하여 투자자가 실질적으로 활용할 수 있는\n"
        "전문적이고 통찰력 있는 한국어 분석 의견을 작성하십시오.\n\n"
        "【출력 형식 규칙 — 반드시 준수】\n"
        "- 정확히 10줄 작성\n"
        "- 각 줄은 ①②③④⑤⑥⑦⑧⑨⑩ 중 하나로 시작\n"
        "- 각 의견은 구체적 수치를 인용하며 1~2문장\n"
        "- 수익성·안정성·성장성·리스크·종합투자의견을 균형 있게\n"
        "- ⑩은 반드시 종합 투자 의견으로 마무리\n"
        "- 번호 외 머리말·꼬리말 금지\n\n"
        "【재무 데이터 JSON】\n"
        f"{fin_str}\n\n"
        "위 데이터를 바탕으로 ①~⑩ 형식 의견 10줄을 작성하십시오."
    )


def call_claude_cli(prompt: str) -> str:
    """AI 의견 생성: CLI 우선(로컬), Anthropic SDK 폴백(Vercel 서버리스)"""

    # ── 방법 1: Claude CLI (로컬 환경) ──────────────────────────────────────
    cmd = _find_claude_cmd()
    if cmd:
        try:
            proc = subprocess.run(
                cmd + ["--print", "-"],
                input=prompt, capture_output=True, text=True,
                encoding="utf-8", timeout=CLI_TIMEOUT,
            )
            if proc.returncode == 0 and proc.stdout.strip():
                return proc.stdout.strip()
        except Exception:
            pass  # CLI 실패 시 SDK로 폴백

    # ── 방법 2: Anthropic Python SDK (Vercel 서버리스) ──────────────────────
    if ANTHROPIC_KEY:
        try:
            import anthropic
            client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
            msg = client.messages.create(
                model="claude-opus-4-7",
                max_tokens=1024,
                thinking={"type": "adaptive"},
                messages=[{"role": "user", "content": prompt}],
            )
            # 텍스트 블록만 추출
            text = "\n".join(
                b.text for b in msg.content if hasattr(b, "text") and b.text
            ).strip()
            if text:
                return text
        except Exception as e:
            raise RuntimeError(f"Anthropic API 오류: {e}")

    raise RuntimeError(
        "AI 의견 생성 불가: Claude CLI 또는 ANTHROPIC_API_KEY 환경변수가 필요합니다."
    )


# ── Excel 생성 ────────────────────────────────────────────────────────────────
def generate_excel(results: list[dict], opinions: dict[str, list[str]]) -> str:
    wb = openpyxl.Workbook()

    def xfill(h):  return PatternFill("solid", fgColor=h.lstrip("#"))
    def xfont(c="FFFFFF", bold=False, size=11): return Font(color=c.lstrip("#"), bold=bold, size=size)
    def xalign(h="center"): return Alignment(horizontal=h, vertical="center", wrap_text=True)
    _s   = Side(style="thin", color="2E4070")
    xbdr = Border(left=_s, right=_s, top=_s, bottom=_s)
    FD   = xfill("0D1B2A"); FN = xfill("1B2A4A"); FA = xfill("4A90D9")
    FR1  = xfill("0D1B2A"); FR2 = xfill("1E3050")
    FGOLD = xfill("2A2000"); FRED = xfill("2A0000"); FBL = xfill("0A1020")

    def W(ws, r, c, v, fill=None, font=None, align=None, border=None):
        cell = ws.cell(row=r, column=c, value=v)
        if fill:   cell.fill      = fill
        if font:   cell.font      = font
        if align:  cell.alignment = align
        if border: cell.border    = border
        return cell

    ws1 = wb.active
    ws1.title = "재무분석"
    now_str = datetime.now().strftime("%Y-%m-%d")

    # ── Sheet1: 회사별 재무 데이터 ─────────────────────────────────────────
    col_offset = 1
    for res in results:
        corp = res["corp"]
        fin  = res["fin"]
        ratios = res["ratios"]

        ws1.merge_cells(f"{get_column_letter(col_offset)}1:{get_column_letter(col_offset+7)}1")
        W(ws1, 1, col_offset,
          f"DART 재무분석 — {corp['corp_name']}  ({now_str})",
          FD, xfont(bold=True, size=14), xalign())
        ws1.row_dimensions[1].height = 36

        row = 3
        ws1.merge_cells(f"{get_column_letter(col_offset)}{row}:{get_column_letter(col_offset+7)}{row}")
        W(ws1, row, col_offset, "손익·재무 현황 (억원)", FN, xfont("4A90D9", True, 12), xalign("left"))
        ws1.row_dimensions[row].height = 26
        row += 1

        hdrs = ["항목", "2023년", "2024년", "2025년", "YoY 23→24", "YoY 24→25"]
        for ci, h in enumerate(hdrs):
            W(ws1, row, col_offset+ci, h, FA, xfont(bold=True), xalign(), xbdr)
        ws1.row_dimensions[row].height = 22; row += 1

        for mi, m in enumerate(["매출액","영업이익","당기순이익","자산총계","부채총계","자본총계"]):
            fb = FR1 if mi % 2 == 0 else FR2
            nums = [(fin.get(str(yr)) or {}).get(m) for yr in YEARS]
            W(ws1, row, col_offset,   m,            fb, xfont(bold=True), xalign("left"), xbdr)
            for ci, n in enumerate(nums):
                W(ws1, row, col_offset+1+ci, fmt_amt(n), fb, xfont(), xalign("right"), xbdr)
            for j in range(1, len(YEARS)):
                v  = yoy(nums[j], nums[j-1])
                ys = fmt_yoy(v) or "N/A"
                fc = "27AE60" if v is not None and v >= 0 else "E74C3C" if v is not None else "FFFFFF"
                W(ws1, row, col_offset+len(YEARS)+j, ys, fb, xfont(fc), xalign(), xbdr)
            ws1.row_dimensions[row].height = 20; row += 1

        row += 1
        ws1.merge_cells(f"{get_column_letter(col_offset)}{row}:{get_column_letter(col_offset+3)}{row}")
        W(ws1, row, col_offset, "재무비율", FN, xfont("4A90D9", True, 12), xalign("left"))
        ws1.row_dimensions[row].height = 26; row += 1

        for ci, h in enumerate(["비율 항목","2023년","2024년","2025년"]):
            W(ws1, row, col_offset+ci, h, FA, xfont(bold=True), xalign(), xbdr)
        ws1.row_dimensions[row].height = 22; row += 1

        for ri, rk in enumerate(["영업이익률","순이익률","ROE","부채비율"]):
            fb = FR1 if ri % 2 == 0 else FR2
            W(ws1, row, col_offset, rk, fb, xfont(bold=True), xalign("left"), xbdr)
            for ci, yr in enumerate(YEARS):
                rv = (ratios.get(str(yr)) or {}).get(rk)
                fc = "27AE60" if rv is not None and ((rk != "부채비율" and rv >= 0) or (rk == "부채비율" and rv < 100)) else "E74C3C" if rv is not None else "FFFFFF"
                W(ws1, row, col_offset+1+ci, fmt_pct(rv), fb, xfont(fc), xalign(), xbdr)
            ws1.row_dimensions[row].height = 20; row += 1

        for ci in range(8):
            ws1.column_dimensions[get_column_letter(col_offset+ci)].width = 14 if ci > 0 else 18
        col_offset += 10

    # ── Sheet5: AI 분석 의견 ───────────────────────────────────────────────
    for sn in ["손익계산서","재무상태표","재무비율"]:
        ws_tmp = wb.create_sheet(sn)
        ws_tmp.cell(1, 1, f"[{sn}] — Sheet1 참조").fill = FN
        ws_tmp.cell(1, 1).font = xfont("8899AA", size=10)

    ws5 = wb.create_sheet("AI 분석 의견")
    total_cols = len(results)
    ws5.merge_cells(f"A1:{get_column_letter(total_cols*2)}1")
    W(ws5, 1, 1, f"DART AI 재무분석 의견 — {now_str} | powered by Claude CLI",
      FD, xfont(bold=True, size=14), xalign())
    ws5.row_dimensions[1].height = 34

    ws5.merge_cells(f"A2:{get_column_letter(total_cols*2)}2")
    corp_names = ", ".join(r["corp"]["corp_name"] for r in results)
    W(ws5, 2, 1, f"비교 대상: {corp_names}  ·  Claude CLI 자동 생성 (참고용, 투자 권유 아님)",
      FN, xfont("8899AA", size=10), xalign())
    ws5.row_dimensions[2].height = 20

    for ci, res in enumerate(results):
        corp  = res["corp"]
        # 프론트엔드가 corp_code를 키로 전송하므로 corp_code 우선 조회
        ops   = opinions.get(corp["corp_code"]) or opinions.get(corp["corp_name"]) or ["AI 분석 의견 없음"]
        cs    = ci * 2 + 1
        ws5.column_dimensions[get_column_letter(cs)].width   = 60
        ws5.column_dimensions[get_column_letter(cs+1)].width = 3
        W(ws5, 4, cs, f"  {corp['corp_name']}", FA, xfont(bold=True, size=12), xalign("left"), xbdr)
        ws5.row_dimensions[4].height = 26
        for ri, line in enumerate(ops):
            r_idx = 5 + ri
            if "⚠" in line or "손실" in line or "오류" in line:
                fl, fc = FRED,  xfont("FF6B6B")
            elif line.startswith("⑩"):
                fl, fc = FGOLD, xfont("F4C020", bold=True)
            elif any(line.startswith(n) for n in ["①","②","③","④","⑤"]):
                fl, fc = FR1,   xfont()
            elif any(line.startswith(n) for n in ["⑥","⑦","⑧","⑨"]):
                fl, fc = FR2,   xfont("D0E8FF")
            else:
                fl, fc = FBL,   xfont("8899AA", size=10)
            W(ws5, r_idx, cs, line, fl, fc,
              Alignment(horizontal="left", vertical="center", wrap_text=True), xbdr)
            ws5.row_dimensions[r_idx].height = 22

    fname    = f"재무분석_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, fname)
    wb.save(filepath)
    return fname


# ── FastAPI 앱 ────────────────────────────────────────────────────────────────
app = FastAPI(title="DART 재무분석 API", version="9.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Pydantic 모델 ─────────────────────────────────────────────────────────────
class SearchReq(BaseModel):
    name: str

class CorpInfo(BaseModel):
    corp_code:  str
    corp_name:  str
    stock_code: str = ""

class AnalyzeMultiReq(BaseModel):
    companies: list[CorpInfo]

class OpinionReq(BaseModel):
    corp_name: str
    fin:       dict

class ExportReq(BaseModel):
    results:  list[dict]
    opinions: dict[str, list[str]] = {}

# ── 엔드포인트 ────────────────────────────────────────────────────────────────
@app.get("/api/health")
def health():
    return {"status": "ok", "timestamp": datetime.now().isoformat()}


@app.post("/api/search")
def search(req: SearchReq):
    if not req.name.strip():
        raise HTTPException(400, "회사명을 입력하세요.")
    try:
        corps = load_corp_codes()
    except Exception as e:
        raise HTTPException(503, f"기업코드 로드 실패: {e}")
    results = search_corp(corps, req.name.strip())
    if not results:
        return []
    return results[:60]


@app.post("/api/analyze-multi")
def analyze_multi(req: AnalyzeMultiReq):
    if not req.companies:
        raise HTTPException(400, "분석할 회사를 선택하세요.")
    if len(req.companies) > 3:
        raise HTTPException(400, "최대 3개 회사까지 분석 가능합니다.")

    corps = [c.model_dump() for c in req.companies]
    try:
        all_fin = fetch_company_data_parallel(corps)
    except Exception as e:
        raise HTTPException(503, f"DART API 오류: {e}")

    results = []
    for corp in corps:
        fin = all_fin.get(corp["corp_code"], {})
        if all(v is None for v in fin.values()):
            results.append({"corp": corp, "fin": {}, "ratios": {}, "yoy": {}, "error": "재무데이터 없음"})
        else:
            results.append(build_company_result(corp, fin))

    return {"results": results}


@app.post("/api/opinion")
def get_opinion(req: OpinionReq):
    try:
        prompt = build_cli_prompt(req.corp_name, req.fin)
        raw    = call_claude_cli(prompt)
        lines  = [ln.strip() for ln in raw.splitlines() if ln.strip()]
        return {"opinions": lines if lines else ["AI 의견 생성 결과가 비어 있습니다."]}
    except subprocess.TimeoutExpired:
        raise HTTPException(504, f"Claude CLI 응답 시간 초과 ({CLI_TIMEOUT}초)")
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/export")
def export(req: ExportReq):
    if not req.results:
        raise HTTPException(400, "분석 결과가 없습니다.")
    try:
        fname = generate_excel(req.results, req.opinions)
        return {"filename": fname}
    except Exception as e:
        raise HTTPException(500, f"Excel 생성 실패: {e}")


@app.get("/api/download/{filename}")
def download(filename: str):
    # 경로 순회 방지
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(400, "잘못된 파일명")
    filepath = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(404, "파일을 찾을 수 없습니다.")
    return FileResponse(
        filepath,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8002, reload=True)
